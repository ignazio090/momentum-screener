#!/usr/bin/env python3
"""
Momentum Screener — zoekt aandelen met een CONSISTENTE opwaartse trend.
GEEN BELEGGINGSADVIES.
"""

import os, sys, csv, time, json
from datetime import datetime, timedelta
import requests
from openpyxl import load_workbook
from openpyxl.styles import PatternFill, Border, Side

# ============================================================
# INSTELLINGEN
# ============================================================
FINNHUB_KEY      = os.environ.get("FINNHUB_KEY", "")
TELEGRAM_TOKEN   = os.environ.get("TELEGRAM_TOKEN", "")
TELEGRAM_CHAT_ID = os.environ.get("TELEGRAM_CHAT_ID", "")

EXCEL_PATH   = "beleggen_tracker.xlsx"
TICKERS_CSV  = "tickers.csv"
HISTORY_JSON = "results_history.json"
DASHBOARD    = "docs/index.html"

PERIODS = {"1D": 1, "1W": 7, "1M": 30, "6M": 182, "1J": 365}
REQUIRE_ALL_RISING = True
MIN_PER_PERIOD     = 0.00

MAX_HITS_IN_MESSAGE = 10
NEWS_PER_STOCK      = 3
SLEEP_BETWEEN_CALLS = 1.05

BASE = "https://finnhub.io/api/v1"


# ============================================================
# DATA
# ============================================================
def get_json(url, params):
    p = dict(params, token=FINNHUB_KEY)
    r = requests.get(url, params=p, timeout=20)
    if r.status_code == 429:
        raise Exception("Rate limit bereikt (429)")
    if r.status_code == 403:
        raise Exception("Geen toegang (403)")
    if r.status_code != 200:
        raise Exception("HTTP " + str(r.status_code))
    return r.json()


def load_universe():
    if not os.path.exists(TICKERS_CSV):
        sys.exit("Geen " + TICKERS_CSV + " gevonden.")
    out = []
    with open(TICKERS_CSV, newline="") as f:
        for row in csv.DictReader(f):
            t = (row.get("ticker") or "").strip()
            if t:
                out.append(t)
    return out


def pct_changes(ticker):
    now = int(time.time())
    frm = now - 400 * 86400
    candles = get_json(
        BASE + "/stock/candle",
        {"symbol": ticker, "resolution": "D", "from": frm, "to": now}
    )
    if candles.get("s") != "ok" or not candles.get("c"):
        return None
    closes = candles["c"]
    stamps = candles["t"]
    last = closes[-1]
    out = {"price": last}
    for label, days in PERIODS.items():
        target = now - days * 86400
        ref = None
        for ts, c in zip(stamps, closes):
            if ts <= target:
                ref = c
            else:
                break
        if ref and ref > 0:
            out[label] = (last - ref) / ref
        else:
            out[label] = None
    return out


def get_news(ticker):
    to_date = datetime.utcnow().date()
    from_date = to_date - timedelta(days=7)
    try:
        data = get_json(
            BASE + "/company-news",
            {"symbol": ticker, "from": str(from_date), "to": str(to_date)}
        )
    except Exception:
        return []
    items = []
    for a in data[:NEWS_PER_STOCK]:
        items.append({
            "headline": a.get("headline", "")[:100],
            "url": a.get("url", "")
        })
    return items


# ============================================================
# FILTER + SCORE
# ============================================================
def is_hit(ch):
    vals = [v for k, v in ch.items() if k != "price"]
    if any(v is None for v in vals):
        return False
    if REQUIRE_ALL_RISING:
        return all(v >= MIN_PER_PERIOD for v in vals)
    else:
        return any(v >= MIN_PER_PERIOD for v in vals)


def momentum_score(ch):
    vals = [v for k, v in ch.items() if k != "price" and v is not None]
    if vals:
        return sum(vals) / len(vals)
    return -99


# ============================================================
# EXCEL
# ============================================================
def update_excel(hits):
    if not os.path.exists(EXCEL_PATH):
        return
    wb = load_workbook(EXCEL_PATH)
    if "Screener" in wb.sheetnames:
        ws = wb["Screener"]
    else:
        ws = wb.create_sheet("Screener")
    row = 5
    while ws.cell(row=row, column=1).value not in (None, ""):
        row += 1
    today = datetime.now().strftime("%Y-%m-%d")
    thin = Side(style="thin", color="D0D0D0")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    for h in hits:
        c = h["changes"]
        news_url = ""
        if h["news"]:
            news_url = h["news"][0].get("url", "")
        vals = [
            today, h["ticker"], h["ticker"], c["price"],
            c.get("1D"), c.get("1W"), c.get("1M"), c.get("6M"), c.get("1J"),
            "JA", news_url
        ]
        for col, v in enumerate(vals, 1):
            cell = ws.cell(row=row, column=col, value=v)
            cell.border = border
            if col == 4:
                cell.number_format = "#,##0.00"
            if 5 <= col <= 9 and v is not None:
                cell.number_format = "0.0%"
                if v >= 0:
                    cell.fill = PatternFill("solid", start_color="DDEFE2")
                else:
                    cell.fill = PatternFill("solid", start_color="F7DDDD")
        row += 1
    wb.save(EXCEL_PATH)
    print("Excel: " + str(len(hits)) + " kandidaten toegevoegd.")


# ============================================================
# HISTORY
# ============================================================
def load_history():
    if os.path.exists(HISTORY_JSON):
        with open(HISTORY_JSON) as f:
            return json.load(f)
    return {}


def save_history(history, hits):
    today = datetime.now().strftime("%Y-%m-%d")
    today_data = []
    for h in hits:
        today_data.append({
            "ticker": h["ticker"],
            "price": h["changes"]["price"],
            "1D": h["changes"].get("1D"),
            "1W": h["changes"].get("1W"),
            "1M": h["changes"].get("1M"),
            "6M": h["changes"].get("6M"),
            "1J": h["changes"].get("1J"),
            "news": h["news"][:2],
            "score": round(momentum_score(h["changes"]) * 100, 1)
        })
    history[today] = today_data
    cutoff = (datetime.now() - timedelta(days=90)).strftime("%Y-%m-%d")
    history = {k: v for k, v in history.items() if k >= cutoff}
    with open(HISTORY_JSON, "w") as f:
        json.dump(history, f, indent=2)
    return history


# ============================================================
# DASHBOARD
# ============================================================
def generate_dashboard(history):
    os.makedirs("docs", exist_ok=True)
    dates = sorted(history.keys(), reverse=True)
    today = dates[0] if dates else "geen data"
    today_hits = history.get(today, [])

    def fmt(v):
        if v is None:
            return "n/a"
        return ("+%.1f%%" % (v * 100)).replace(".", ",")

    def cls(v):
        if v is None:
            return "muted"
        if v >= 0:
            return "up"
        return "down"

    rows_html = ""
    for s in today_hits:
        news_links = ""
        for n in s.get("news", []):
            url = n.get("url", "")
            title = n.get("headline", "")[:60]
            if url:
                news_links += '<a class="news" href="' + url + '" target="_blank">' + title + '</a><br>'
        if not news_links:
            news_links = '<span class="muted">geen</span>'
        rows_html += "<tr>"
        rows_html += '<td><span class="tk">' + s["ticker"] + "</span></td>"
        rows_html += "<td>$" + ("%.2f" % s["price"]) + "</td>"
        rows_html += '<td class="' + cls(s.get("1D")) + '">' + fmt(s.get("1D")) + "</td>"
        rows_html += '<td class="' + cls(s.get("1W")) + '">' + fmt(s.get("1W")) + "</td>"
        rows_html += '<td class="' + cls(s.get("1M")) + '">' + fmt(s.get("1M")) + "</td>"
        rows_html += '<td class="' + cls(s.get("6M")) + '">' + fmt(s.get("6M")) + "</td>"
        rows_html += '<td class="' + cls(s.get("1J")) + '">' + fmt(s.get("1J")) + "</td>"
        rows_html += '<td class="score">' + str(s.get("score", "")) + "</td>"
        rows_html += "<td>" + news_links + "</td>"
        rows_html += "</tr>"

    if not rows_html:
        rows_html = '<tr><td colspan="9" class="empty">Geen kandidaten vandaag.</td></tr>'

    hist_html = ""
    for d in dates[:30]:
        n = len(history[d])
        tickers = ", ".join(h["ticker"] for h in history[d][:8])
        if len(history[d]) > 8:
            tickers += " +" + str(len(history[d]) - 8)
        hist_html += '<div class="hist-row">'
        hist_html += '<span class="hist-date">' + d + "</span>"
        hist_html += '<span class="hist-count">' + str(n) + "</span>"
        hist_html += '<span class="hist-tickers">' + tickers + "</span>"
        hist_html += "</div>"

    if not hist_html:
        hist_html = '<div class="hist-row"><span class="muted">Nog geen historie.</span></div>'

    p = []
    p.append("<!DOCTYPE html>")
    p.append("<html lang='nl'><head><meta charset='UTF-8'>")
    p.append("<meta name='viewport' content='width=device-width,initial-scale=1'>")
    p.append("<title>Momentum Screener</title>")
    p.append("<style>")
    p.append("@import url('https://fonts.googleapis.com/css2?family=Space+Grotesk:wght@500;600;700&family=Inter:wght@400;500;600&family=JetBrains+Mono:wght@400;500;700&display=swap');")
    p.append(":root{--bg:#0B0E0C;--panel:#141A16;--panel2:#1B231D;--line:#26302A;--text:#E8EFE9;--muted:#8A958C;--up:#4ADE80;--down:#F87171;--accent:#4ADE80;}")
    p.append("*{box-sizing:border-box;margin:0;padding:0;}")
    p.append("body{background:var(--bg);color:var(--text);font-family:Inter,sans-serif;padding:36px 20px 80px;}")
    p.append(".wrap{max-width:1120px;margin:0 auto;}")
    p.append(".eyebrow{font-family:JetBrains Mono,monospace;font-size:11px;letter-spacing:.18em;text-transform:uppercase;color:var(--accent);}")
    p.append("h1{font-family:Space Grotesk,sans-serif;font-weight:700;font-size:clamp(28px,4vw,42px);letter-spacing:-.02em;margin:10px 0;}")
    p.append(".sub{color:var(--muted);font-size:14px;max-width:600px;line-height:1.55;margin-bottom:6px;}")
    p.append(".badge{display:inline-flex;padding:5px 11px;border:1px solid var(--line);border-radius:8px;font-family:JetBrains Mono,monospace;font-size:11px;color:var(--muted);margin-bottom:28px;}")
    p.append(".panel{background:var(--panel);border:1px solid var(--line);border-radius:14px;overflow:hidden;margin-bottom:24px;}")
    p.append(".panel-head{padding:16px 20px;border-bottom:1px solid var(--line);display:flex;justify-content:space-between;align-items:center;flex-wrap:wrap;gap:8px;}")
    p.append(".panel-head h2{font-family:Space Grotesk;font-size:16px;font-weight:600;}")
    p.append(".panel-head .meta{font-family:JetBrains Mono;font-size:10.5px;color:var(--muted);}")
    p.append("table{width:100%;border-collapse:collapse;}")
    p.append("th{font-family:JetBrains Mono,monospace;font-size:10px;letter-spacing:.06em;text-transform:uppercase;color:var(--muted);font-weight:500;text-align:right;padding:12px;border-bottom:1px solid var(--line);}")
    p.append("th:first-child,td:first-child{text-align:left;padding-left:20px;}")
    p.append("td{padding:12px;text-align:right;font-family:JetBrains Mono,monospace;font-size:12.5px;border-bottom:1px solid var(--line);vertical-align:top;}")
    p.append("tr:last-child td{border-bottom:none;}")
    p.append("tbody tr:hover{background:var(--panel2);}")
    p.append(".tk{font-family:Space Grotesk;font-weight:700;font-size:14px;}")
    p.append(".up{color:var(--up);} .down{color:var(--down);} .muted{color:var(--muted);}")
    p.append(".score{color:var(--accent);font-weight:700;}")
    p.append(".news{color:var(--accent);text-decoration:none;font-family:Inter;font-size:11px;display:inline-block;margin:2px 0;}")
    p.append(".news:hover{text-decoration:underline;}")
    p.append(".empty{text-align:center;color:var(--muted);font-family:Inter;font-size:13px;padding:30px;}")
    p.append(".hist-row{display:flex;gap:14px;align-items:baseline;padding:9px 20px;border-bottom:1px solid var(--line);font-size:13px;}")
    p.append(".hist-row:last-child{border-bottom:none;}")
    p.append(".hist-date{font-family:JetBrains Mono;font-size:12px;color:var(--muted);min-width:90px;}")
    p.append(".hist-count{font-family:Space Grotesk;font-weight:700;color:var(--accent);min-width:30px;}")
    p.append(".hist-tickers{color:var(--text);font-size:12.5px;}")
    p.append("footer{margin-top:36px;padding-top:16px;border-top:1px solid var(--line);font-size:10.5px;color:var(--muted);font-family:JetBrains Mono,monospace;line-height:1.6;}")
    p.append("</style></head><body>")
    p.append("<div class='wrap'>")
    p.append("<div class='eyebrow'>MOMENTUM SCREENER</div>")
    p.append("<h1>Consistente stijgers</h1>")
    p.append("<p class='sub'>Aandelen die over dag, week, maand, halfjaar en jaar tegelijk stijgen.</p>")
    p.append("<div class='badge'>Laatst bijgewerkt: " + today + " | " + str(len(today_hits)) + " kandidaten | 319 aandelen gescand</div>")
    p.append("<div class='panel'>")
    p.append("<div class='panel-head'><h2>Kandidaten - " + today + "</h2><span class='meta'>FILTER: POSITIEF OVER ALLE PERIODES</span></div>")
    p.append("<table><thead><tr>")
    p.append("<th>Ticker</th><th>Koers</th><th>1D</th><th>1W</th><th>1M</th><th>6M</th><th>1J</th><th>Score</th><th>Nieuws</th>")
    p.append("</tr></thead><tbody>")
    p.append(rows_html)
    p.append("</tbody></table></div>")
    p.append("<div class='panel'>")
    p.append("<div class='panel-head'><h2>Historie (laatste 30 dagen)</h2><span class='meta'>AANTAL KANDIDATEN PER DAG</span></div>")
    p.append(hist_html)
    p.append("</div>")
    p.append("<footer>MOMENTUM SCREENER - AUTOMATISCH GEGENEREERD - DIT IS GEEN BELEGGINGSADVIES</footer>")
    p.append("</div></body></html>")

    with open(DASHBOARD, "w") as f:
        f.write("\n".join(p))
    print("Dashboard gegenereerd: " + DASHBOARD)


# ============================================================
# TELEGRAM
# ============================================================
def fmt_pct(v):
    if v is None:
        return "n/a"
    return ("+%.1f%%" % (v * 100)).replace(".", ",")


def build_message(hits):
    if not hits:
        return "Geen aandelen met een consistente uptrend over alle periodes vandaag."
    lines = []
    lines.append("Consistente stijgers - " + datetime.now().strftime("%d-%m-%Y"))
    lines.append(str(len(hits)) + " aandelen stijgen over ALLE periodes:")
    lines.append("")
    for h in hits[:MAX_HITS_IN_MESSAGE]:
        c = h["changes"]
        parts = []
        for k in PERIODS:
            parts.append(k + " " + fmt_pct(c.get(k)))
        trend = "  ".join(parts)
        lines.append("* " + h["ticker"])
        lines.append("   " + trend)
        for n in h["news"][:1]:
            if n.get("url"):
                lines.append("   " + n["headline"][:50])
                lines.append("   " + n["url"])
    if len(hits) > MAX_HITS_IN_MESSAGE:
        lines.append("")
        lines.append("+" + str(len(hits) - MAX_HITS_IN_MESSAGE) + " meer - zie dashboard.")
    lines.append("")
    lines.append("Geen advies - zelf onderzoeken.")
    return "\n".join(lines)


def send_telegram(text):
    if not TELEGRAM_TOKEN or not TELEGRAM_CHAT_ID:
        print("Telegram niet geconfigureerd.")
        print(text)
        return
    url = "https://api.telegram.org/bot" + TELEGRAM_TOKEN + "/sendMessage"
    r = requests.post(url, data={
        "chat_id": TELEGRAM_CHAT_ID,
        "text": text,
        "disable_web_page_preview": True
    }, timeout=20)
    if r.ok:
        print("Telegram verstuurd.")
    else:
        print("Telegram-fout: " + r.text)


# ============================================================
# MAIN
# ============================================================
def run_real():
    if not FINNHUB_KEY:
        sys.exit("Stel FINNHUB_KEY in als omgevingsvariabele.")

    key_len = len(FINNHUB_KEY)
    print("API-key lengte: " + str(key_len) + " tekens (eerste 4: " + FINNHUB_KEY[:4] + "...)")
    if key_len < 10:
        print("Key lijkt te kort of leeg.")
        sys.exit(1)

    # Test quote endpoint
    try:
        r = requests.get(BASE + "/quote",
                         params={"symbol": "AAPL", "token": FINNHUB_KEY},
                         timeout=20)
        print("Test /quote: HTTP " + str(r.status_code))
        print("Response: " + r.text[:200])
    except Exception as e:
        print("Test /quote mislukt: " + str(e))

    # Test candle endpoint
    try:
        r2 = requests.get(BASE + "/stock/candle",
                          params={"symbol": "AAPL", "resolution": "D",
                                  "from": int(time.time()) - 7 * 86400,
                                  "to": int(time.time()),
                                  "token": FINNHUB_KEY},
                          timeout=20)
        print("Test /candle: HTTP " + str(r2.status_code))
        print("Response: " + r2.text[:200])
        if r2.status_code != 200:
            print("Candle endpoint werkt niet. Stoppen.")
            sys.exit(1)
    except Exception as e:
        print("Test /candle mislukt: " + str(e))
        sys.exit(1)

    print("")
    print("API werkt! Start screenen...")
    universe = load_universe()
    print("Screenen van " + str(len(universe)) + " aandelen...")

    hits = []
    errors = 0
    for i, tk in enumerate(universe, 1):
        try:
            ch = pct_changes(tk)
        except Exception as e:
            errors += 1
            if errors <= 5:
                print("  x " + tk + ": " + str(e))
            elif errors == 6:
                print("  ... verdere fouten niet getoond")
            time.sleep(SLEEP_BETWEEN_CALLS)
            continue
        if ch and is_hit(ch):
            news = get_news(tk)
            hits.append({"ticker": tk, "changes": ch, "news": news})
            print("  v " + tk)
        time.sleep(SLEEP_BETWEEN_CALLS)
        if i % 50 == 0:
            print("  ..." + str(i) + "/" + str(len(universe)) + " (" + str(len(hits)) + " hits, " + str(errors) + " fouten)")

    hits.sort(key=lambda h: momentum_score(h["changes"]), reverse=True)
    print("")
    print("Klaar: " + str(len(hits)) + " kandidaten uit " + str(len(universe)) + " aandelen (" + str(errors) + " fouten).")
    print("")

    update_excel(hits)
    history = load_history()
    history = save_history(history, hits)
    generate_dashboard(history)
    send_telegram(build_message(hits))


def run_demo():
    print("DEMO-modus.")
    print("")
    demo = [
        {"ticker": "NVDA",
         "changes": {"price": 118.5, "1D": 0.021, "1W": 0.034, "1M": 0.071, "6M": 0.18, "1J": 0.42},
         "news": [{"headline": "NVIDIA kondigt nieuwe chip aan", "url": "https://example.com/nvda"}]},
        {"ticker": "ADYEN",
         "changes": {"price": 1450, "1D": 0.012, "1W": 0.028, "1M": 0.04, "6M": 0.10, "1J": 0.22},
         "news": [{"headline": "Adyen breidt uit", "url": "https://example.com/adyen"}]},
        {"ticker": "SHELL",
         "changes": {"price": 32.05, "1D": -0.004, "1W": 0.011, "1M": 0.03, "6M": 0.06, "1J": 0.14},
         "news": []},
    ]
    hits = [h for h in demo if is_hit(h["changes"])]
    hits.sort(key=lambda h: momentum_score(h["changes"]), reverse=True)
    print(str(len(hits)) + " van " + str(len(demo)) + " voldoen aan EN-filter.")
    print("")
    update_excel(hits)
    history = load_history()
    history = save_history(history, hits)
    generate_dashboard(history)
    print("")
    print("--- Telegram-bericht ---")
    print("")
    print(build_message(hits))


if __name__ == "__main__":
    if "--demo" in sys.argv:
        run_demo()
    else:
        run_real()
